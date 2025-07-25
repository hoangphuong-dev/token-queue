# -*- coding: utf-8 -*-
from odoo import models, api, fields, _
import logging
import re
from odoo.exceptions import ValidationError

MATE_HMS_APPOINTMENT = 'mate_hms.appointment'
MATE_HSM_PACKAGE = 'mate_hms.package'
MATE_HMS_HANDLE_PACKAGE = 'mate_hms.handle.package'
MATE_HMS_HANDLE_UPLOAD_PACKAGE = 'mate_hms_subscriptions.handle.upload.package'
MATE_HMS_HANDLE_PACKAGE_LINE = 'mate_hms.handle.package.line'
MATE_HMS_SUBSCRIPTION_CATEGORY = 'mate_hms.subscription.category'
MATE_HSM_PACKAGE_LINE = 'mate_hms.package.line'

_logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except ImportError:
    _logger.warning("Please install openpyxl to handle Excel files.")
    load_workbook = None


class MateHmsHandlePackageLine(models.TransientModel):
    _name = MATE_HMS_HANDLE_PACKAGE_LINE

    code = fields.Char(string='Code')
    name = fields.Char(string='Name')
    quantity = fields.Integer(string='Quantity')
    handle_package_id = fields.Many2one(MATE_HMS_HANDLE_UPLOAD_PACKAGE, string='Handle Upload Package')
    package_name = fields.Char(string='Name Package')
    start_date = fields.Date("Start Date")
    end_date = fields.Date("End Date")
    category = fields.Char(string='Category')
    package_code = fields.Char(string='Package Code')


class MateHandleUploadPackage(models.TransientModel):
    _name = MATE_HMS_HANDLE_UPLOAD_PACKAGE
    _inherit = ['mate.excel.handler.base']
    _description = "Wizard to upload and handle consumed services from Excel file"

    excel_file = fields.Binary()
    excel_file_name = fields.Char()
    package_line_ids = fields.One2many(MATE_HMS_HANDLE_PACKAGE_LINE, 'handle_package_id')

    @api.onchange('excel_file')
    def onchange_excel_file(self):
        if self.excel_file and load_workbook:
            self._validate_file_name(self.excel_file_name)
            header, data_rows = self._get_header_and_data(self.excel_file)
            if not data_rows:
                raise ValidationError(_("No data found in the uploaded file."))
            header = tuple(re.sub(r"\(\*\)\s*\n", "", h).strip().replace('\n(yyyy-mm-dd)', '') for h in header)
            try:
                data_rows = sorted(data_rows, key=lambda x: x[0])
                # Gọi hàm kiểm tra tồn tại
                self._check_existing_package_codes(header, data_rows)
                list_package_line = self._process_excel_rows(header, data_rows)
                self.package_line_ids = [(0, 0, line) for line in list_package_line]
            except Exception:
                raise ValidationError(_("An error occurred while processing the Excel file. Please check the file format and data."))

    def _check_existing_package_codes(self, header, data_rows):
        """Raise UserError if any package_code already exists in DB"""
        code_line_map = {}
        for index, row in enumerate(data_rows, start=2):  # dòng 2 trở đi là data
            try:
                code = row[header.index('Mã gói')].strip()
                if code:
                    code_line_map.setdefault(code, []).append(index)
            except Exception:
                continue

        if not code_line_map:
            return

        existing_packages = self.env[MATE_HSM_PACKAGE].search([
            ('code', 'in', list(code_line_map.keys()))
        ])

        if existing_packages:
            raise ValidationError(_("The following packages already exist in the system:"))

    def _process_excel_rows(self, header, data_rows):
        list_package_line = []
        for row_index, row in enumerate(data_rows, start=2):
            data = dict(zip(header, row))
            list_package_line.append({
                'code': data.get('Mã dịch vụ'),
                'name': data.get('Tên dịch vụ'),
                'quantity': data.get('Số lượng'),
                'package_name': data.get('Tên gói'),
                'start_date': self._parse_request_time(data.get('Ngày bắt đầu có hiệu lực')),
                'end_date': self._parse_request_time(data.get('Ngày hết hiệu lực')),
                'category': data.get('Loại gói'),
                'package_code': data.get('Mã gói'),
            })
        return list_package_line

    def _prepare_caches(self):
        """Prepare category, product and package caches from package lines"""
        category_model = self.env[MATE_HMS_SUBSCRIPTION_CATEGORY]
        product_model = self.env['product.product']
        package_model = self.env[MATE_HSM_PACKAGE]

        # Collect unique values
        category_names = {line.category or '' for line in self.package_line_ids}
        product_codes = {line.code or '' for line in self.package_line_ids}
        package_codes = {line.package_code or '' for line in self.package_line_ids}

        # Create caches
        category_cache = {
            rec.name: rec for rec in category_model.search([('name', 'in', list(category_names))])
        }
        product_cache = {
            rec.default_code: rec for rec in product_model.search([('default_code', 'in', list(product_codes))])
        }
        package_cache = {
            rec.code: rec for rec in package_model.search([('code', 'in', list(package_codes))])
        }

        return category_cache, product_cache, package_cache, category_model, product_model

    def _process_package_lines(self, category_cache, product_cache, package_cache, category_model, product_model):
        """Process package lines and return package line data"""
        package_model = self.env[MATE_HSM_PACKAGE]
        package_line_data = []

        for line in self.package_line_ids:
            # Get or create category
            category_name = line.category or ''
            if category_name not in category_cache:
                category = category_model.create({'name': category_name})
                category_cache[category_name] = category
            else:
                category = category_cache[category_name]

            # Get or create product
            product_code = line.code or ''
            if product_code not in product_cache:
                product = product_model.create({
                    'name': line.name,
                    'default_code': product_code,
                    'type': 'service',
                })
                product_cache[product_code] = product
            else:
                product = product_cache[product_code]

            # Get or create package
            package_code = line.package_code or ''
            if package_code not in package_cache:
                package = package_model.create({
                    'name': line.package_name,
                    'code': package_code,
                    'start_date': line.start_date or fields.Date.today(),
                    'end_date': line.end_date,
                    'category_id': category.id,
                })
                package_cache[package_code] = package
            else:
                package = package_cache[package_code]

            # Prepare package line data
            package_line_data.append({
                'name': line.name,
                'product_uom_qty': line.quantity,
                'order_id': package.id,
                'product_id': product.id,
            })

        return package_line_data

    def save_package(self):
        # Get caches and models
        category_cache, product_cache, package_cache, category_model, product_model = self._prepare_caches()

        # Process package lines
        package_line_data = self._process_package_lines(
            category_cache, product_cache, package_cache, category_model, product_model
        )

        # Create all package lines at once
        self.env[MATE_HSM_PACKAGE_LINE].create(package_line_data)

        action = self.env.ref('mate_hms_subscriptions.mate_hms_package_action').read()[0]
        return action

    def download_template_file(self):
        """Download the template file"""
        return {
            'type': 'ir.actions.act_url',
            'url': '/mate_hms_subscriptions/static/description/sample_package.xlsx',
            'target': 'self',
        }
